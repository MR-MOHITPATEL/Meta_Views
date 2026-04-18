"""
Custom View Builder — query Raw Dump with any dimension/metric/date combination.

Flow:
  1. Load Raw Dump (via sheets_loader)
  2. normalize_raw_dump() — standard column names + type coercion
  3. Optional: explode_pincodes() — only when Pincode is a selected dimension
  4. Apply date + entity filters
  5. Group by selected dimensions, aggregate selected metrics
  6. Recompute ratio metrics after aggregation
  7. Return DataFrame + attach image_url / ad_id when Ad name is a dimension
"""

from __future__ import annotations

import io
import json
import logging
import os
from pathlib import Path

import pandas as pd

logger = logging.getLogger(__name__)

# ── Available dimensions and metrics ──────────────────────────────────────────

ALL_DIMENSIONS = ["Date", "Campaign name", "Ad set name", "Ad name", "Pincode"]

ALL_METRICS = [
    "Spend", "Impressions", "Clicks", "Purchases", "Revenue",
    "Add to Cart", "Landing Page Views",
    "CTR", "CPC", "CPM", "CPT", "C2V Ratio", "CVR", "ROAS",
    "Pincode Days",
]

# Metrics that must be SUMMED (additive)
_ADDITIVE = ["Spend", "Impressions", "Clicks", "Purchases", "Revenue",
             "Add to Cart", "Landing Page Views"]

# Passthrough columns — taken directly from the raw dump (not summed, not recomputed)
# These are pre-computed by Meta and carried through via a weighted-mean proxy
_PASSTHROUGH = ["C2V Ratio"]

# Metrics that are RECOMPUTED after summing additive cols
_DERIVED = ["CTR", "CPC", "CPM", "CPT", "CVR", "ROAS"]

# Metrics computed from pincode explosion (not additive, not ratio)
_PINCODE_METRICS = ["Pincode Days"]

# Metrics displayed as percentages (stored as decimals, shown ×100)
PCT_METRICS = {"CTR", "CVR", "C2V Ratio"}

# Default metric set shown in UI
DEFAULT_METRICS = ["Spend", "Purchases", "Clicks", "Impressions", "Revenue",
                   "CTR", "CPC", "CPT", "C2V Ratio", "ROAS"]

# ── Saved view config persistence ─────────────────────────────────────────────

_CONFIGS_FILE = Path(__file__).parent / "saved_views.json"


def load_saved_configs() -> dict[str, dict]:
    """Load all saved view configs from disk."""
    if _CONFIGS_FILE.exists():
        try:
            return json.loads(_CONFIGS_FILE.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def save_config(name: str, config: dict) -> None:
    """Persist a named view config to disk."""
    configs = load_saved_configs()
    configs[name] = config
    _CONFIGS_FILE.write_text(
        json.dumps(configs, indent=2, ensure_ascii=False), encoding="utf-8"
    )


def delete_config(name: str) -> None:
    configs = load_saved_configs()
    configs.pop(name, None)
    _CONFIGS_FILE.write_text(
        json.dumps(configs, indent=2, ensure_ascii=False), encoding="utf-8"
    )


# ── Core aggregation ──────────────────────────────────────────────────────────

def _recompute_ratios(df: pd.DataFrame) -> pd.DataFrame:
    """Add derived ratio columns after additive aggregation."""
    c = df.get("Clicks",      pd.Series(0, index=df.index))
    i = df.get("Impressions", pd.Series(0, index=df.index))
    s = df.get("Spend",       pd.Series(0, index=df.index))
    p = df.get("Purchases",   pd.Series(0, index=df.index))
    r = df.get("Revenue",     pd.Series(0, index=df.index))

    if "CTR"  in df.columns: df["CTR"]  = (c / i.replace(0, float("nan"))).round(4)
    if "CPC"  in df.columns: df["CPC"]  = (s / c.replace(0, float("nan"))).round(2)
    if "CPM"  in df.columns: df["CPM"]  = (s / i.replace(0, float("nan")) * 1000).round(2)
    if "CPT"  in df.columns: df["CPT"]  = (s / p.replace(0, float("nan"))).round(2)
    if "CVR"  in df.columns: df["CVR"]  = (p / c.replace(0, float("nan"))).round(4)
    if "ROAS" in df.columns: df["ROAS"] = (r / s.replace(0, float("nan"))).round(2)
    return df


def build_custom_view(
    raw_df: pd.DataFrame,
    dimensions: list[str],
    metrics: list[str],
    date_from: str | None = None,
    date_to:   str | None = None,
    filter_campaigns: list[str] | None = None,
    filter_adsets:    list[str] | None = None,
    filter_ads:       list[str] | None = None,
    cpt_min:     float | None = None,
    cpt_max:     float | None = None,
    ctr_min:     float | None = None,
    ctr_max:     float | None = None,
    revenue_min: float | None = None,
    revenue_max: float | None = None,
) -> pd.DataFrame:
    """
    Build a custom view from raw_df with the given dimensions and metrics.

    Parameters
    ----------
    raw_df          : Raw Dump DataFrame (already loaded from Sheets)
    dimensions      : List of columns to group by (e.g. ["Campaign name", "Ad name"])
    metrics         : List of metric columns to include
    date_from/to    : Optional ISO date strings for filtering
    filter_*        : Optional entity filters (OR match within each list)
    """
    from view_builder import normalize_raw_dump, explode_pincodes, _extract_pincodes

    if raw_df.empty:
        return pd.DataFrame()

    # Step 1: Normalize — this is the metric_df (pre-explosion, used for summing spend etc.)
    metric_df = normalize_raw_dump(raw_df)

    _want_pincodes = "Pincode" in dimensions
    _has_pincode_col = "Pincode" in metric_df.columns

    # Step 2: Date filter on metric_df
    if "Date" in metric_df.columns:
        metric_df["Date"] = pd.to_datetime(metric_df["Date"], errors="coerce")
        if date_from:
            metric_df = metric_df[metric_df["Date"] >= pd.Timestamp(date_from)]
        if date_to:
            metric_df = metric_df[metric_df["Date"] <= pd.Timestamp(date_to)]

    # Step 3: Entity filters on metric_df (OR match)
    if filter_campaigns and "Campaign name" in metric_df.columns:
        mask = pd.Series(False, index=metric_df.index)
        for c in filter_campaigns:
            mask |= metric_df["Campaign name"].astype(str).str.lower().str.contains(
                str(c).lower(), na=False)
        metric_df = metric_df[mask]

    if filter_adsets and "Ad set name" in metric_df.columns:
        mask = pd.Series(False, index=metric_df.index)
        for a in filter_adsets:
            mask |= metric_df["Ad set name"].astype(str).str.lower().str.contains(
                str(a).lower(), na=False)
        metric_df = metric_df[mask]

    if filter_ads and "Ad name" in metric_df.columns:
        mask = pd.Series(False, index=metric_df.index)
        for a in filter_ads:
            mask |= metric_df["Ad name"].astype(str).str.lower().str.contains(
                str(a).lower(), na=False)
        metric_df = metric_df[mask]

    if metric_df.empty:
        return pd.DataFrame()

    # Step 4: Build pc_df — exploded version used ONLY for collecting unique pincodes per group.
    # Metrics are NEVER summed on pc_df to avoid inflation (a row with 2 pincodes would
    # double-count spend if we summed after explosion).
    pc_df = explode_pincodes(metric_df.copy()) if _has_pincode_col else metric_df.copy()
    _pincode_col = "Pincode" if "Pincode" in pc_df.columns else None

    # Step 5: Determine aggregation columns
    # Never group BY individual Pincode — pincodes are collected as a comma-separated column.
    dim_cols     = [d for d in dimensions if d in metric_df.columns and d != "Pincode"]
    add_cols     = [m for m in metrics if m in _ADDITIVE and m in metric_df.columns]
    derived_cols = [m for m in metrics if m in _DERIVED]
    pass_cols    = [m for m in metrics if m in _PASSTHROUGH and m in metric_df.columns]

    if not dim_cols:
        # Only "Pincode" selected — return grand total with all pincodes listed
        if _pincode_col and not pc_df.empty:
            all_pincodes = ", ".join(sorted(set(
                str(v) for v in pc_df[_pincode_col].dropna()
                if str(v).strip() and str(v) != "nan"
            )))
            result = pd.DataFrame({"Pincodes": [all_pincodes]})
            for m in add_cols:
                result[m] = metric_df[m].sum()
            return result
        return pd.DataFrame()

    # Pull in extra additive cols needed to compute derived ratios
    _needed_for_ratios = {
        "CTR":  ["Clicks", "Impressions"],
        "CPC":  ["Spend",  "Clicks"],
        "CPM":  ["Spend",  "Impressions"],
        "CPT":  ["Spend",  "Purchases"],
        "CVR":  ["Purchases", "Clicks"],
        "ROAS": ["Revenue", "Spend"],
    }
    extra_additive = set()
    for d in derived_cols:
        for dep in _needed_for_ratios.get(d, []):
            if dep in metric_df.columns and dep not in add_cols:
                extra_additive.add(dep)
    all_additive = list(dict.fromkeys(add_cols + list(extra_additive)))

    # Step 6: Aggregate metrics on pre-explosion metric_df (no inflation)
    _dim_cols_present = [c for c in dim_cols if c in metric_df.columns]
    if all_additive:
        agg = metric_df.groupby(_dim_cols_present, dropna=False)[all_additive].sum().reset_index()
    else:
        agg = metric_df[_dim_cols_present].drop_duplicates().reset_index(drop=True)

    # Passthrough metrics (C2V Ratio) as weighted mean
    for pc in pass_cols:
        if pc in metric_df.columns:
            _pt = (
                metric_df.groupby(_dim_cols_present, dropna=False)[pc]
                .mean()
                .reset_index()
            )
            agg = agg.merge(_pt, on=_dim_cols_present, how="left")

    # Step 7: Collect comma-joined Pincodes from pc_df (explosion is safe here — no summing)
    if _pincode_col and _dim_cols_present:
        _pc_dim_cols = [c for c in _dim_cols_present if c in pc_df.columns]
        if _pc_dim_cols:
            _pc_series = (
                pc_df.groupby(_pc_dim_cols, dropna=False)[_pincode_col]
                .apply(lambda s: ", ".join(
                    sorted(set(str(v) for v in s if pd.notna(v) and str(v).strip() and str(v) != "nan"))
                ))
                .reset_index()
                .rename(columns={_pincode_col: "Pincodes"})
            )
            agg = agg.merge(_pc_series, on=_pc_dim_cols, how="left")
            if "Pincodes" in agg.columns:
                agg["Pincodes"] = agg["Pincodes"].fillna("")

    # Step 8: Recompute ratio metrics after aggregation
    for d in derived_cols:
        agg[d] = float("nan")
    agg = _recompute_ratios(agg)

    # Step 9: Build final column list
    final_cols = _dim_cols_present + [m for m in metrics if m in agg.columns and m != "Pincode"]

    # Step 9a: Compute Pincode Days when requested
    # Logic: COUNT DISTINCT (Date, Campaign name, Pincode) grouped by dim_cols.
    # Uses pc_df (exploded) so each pincode is its own row; dedup key is always
    # (Date, Campaign name, Pincode) regardless of what dimensions are selected.
    if "Pincode Days" in metrics and _pincode_col and _dim_cols_present:
        from view_builder import _pincode_day_per_group
        _pc_dim_for_pd = [c for c in _dim_cols_present if c in pc_df.columns]
        if _pc_dim_for_pd:
            _pd_series = _pincode_day_per_group(pc_df, _pc_dim_for_pd)
            agg = agg.merge(_pd_series, on=_pc_dim_for_pd, how="left")
            agg["pincode_day"] = agg["pincode_day"].fillna(0).astype(int)
            agg = agg.rename(columns={"pincode_day": "Pincode Days"})

    # Step 9b: Attach image_url + ad_id when Ad name is a dimension
    for extra in ("image_url", "ad_id"):
        if "Ad name" in _dim_cols_present and extra in metric_df.columns:
            url_map = (
                metric_df[["Ad name", extra]]
                .replace("", pd.NA)
                .dropna(subset=[extra])
                .drop_duplicates("Ad name")
                .set_index("Ad name")[extra]
            )
            agg[extra] = agg["Ad name"].map(url_map).fillna("")
            if extra not in final_cols:
                final_cols.append(extra)

    # Step 9c: Build final column list in user-specified order
    # Dimensions first (in the order the user selected them), then metrics in order.
    # "Pincode" dimension maps to "Pincodes" output column; insert it at its original position.
    final_cols = []
    for d in dimensions:
        if d == "Pincode":
            if "Pincodes" in agg.columns and (_want_pincodes or _has_pincode_col):
                final_cols.append("Pincodes")
        elif d in agg.columns:
            final_cols.append(d)

    # Metrics in user-selected order (skip Pincode, skip already-added dims)
    _already = set(final_cols)
    for m in metrics:
        if m != "Pincode" and m in agg.columns and m not in _already:
            final_cols.append(m)
            _already.add(m)

    # Re-add image_url / ad_id after metrics rebuild
    for extra in ("image_url", "ad_id"):
        if extra in agg.columns and extra not in final_cols:
            if "Ad name" in _dim_cols_present:
                final_cols.append(extra)

    agg = agg[[c for c in final_cols if c in agg.columns]]

    # Step 10: Apply metric threshold filters (post-aggregation)
    # CTR is stored as decimal (e.g. 0.05 = 5%); user inputs are in % so divide by 100.
    if "CTR" in agg.columns:
        if ctr_min is not None:
            agg = agg[agg["CTR"] >= ctr_min / 100]
        if ctr_max is not None:
            agg = agg[agg["CTR"] <= ctr_max / 100]
    if "CPT" in agg.columns:
        if cpt_min is not None:
            agg = agg[agg["CPT"] >= cpt_min]
        if cpt_max is not None:
            agg = agg[agg["CPT"] <= cpt_max]
    if "Revenue" in agg.columns:
        if revenue_min is not None:
            agg = agg[agg["Revenue"] >= revenue_min]
        if revenue_max is not None:
            agg = agg[agg["Revenue"] <= revenue_max]

    # Format Date column
    if "Date" in agg.columns:
        agg["Date"] = pd.to_datetime(agg["Date"]).dt.strftime("%Y-%m-%d")
        agg = agg.sort_values("Date", ascending=False).reset_index(drop=True)

    return agg.reset_index(drop=True)


# ── Excel export with full formatting ────────────────────────────────────────

def build_formatted_excel(df: pd.DataFrame, view_name: str = "Custom View") -> bytes:
    """
    Export DataFrame to a formatted Excel file:
      - Bold white headers on dark teal background
      - Alternating light/white row shading
      - Auto-width columns
      - Embedded ad images when image_url column is present
    """
    try:
        import requests as _req
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from PIL import Image as PILImage
    except ImportError as e:
        logger.warning(f"Excel formatting unavailable: {e}")
        buf = io.BytesIO()
        df.to_excel(buf, index=False)
        return buf.getvalue()

    # Convert percentage metrics (stored as decimals) → display values ×100
    df = df.copy()
    for _pct_col in PCT_METRICS:
        if _pct_col in df.columns:
            df[_pct_col] = pd.to_numeric(df[_pct_col], errors="coerce") * 100

    has_images = "image_url" in df.columns and \
        df["image_url"].astype(str).str.startswith("http").any()

    # Image byte cache: url → PNG bytes
    _img_cache: dict[str, bytes | None] = {}

    def _fetch(url: str) -> bytes | None:
        if not url or not str(url).startswith("http"):
            return None
        if url in _img_cache:
            return _img_cache[url]
        try:
            r = _req.get(url, timeout=8)
            r.raise_for_status()
            img = PILImage.open(io.BytesIO(r.content))
            img.thumbnail((120, 120), PILImage.LANCZOS)
            out = io.BytesIO()
            img.save(out, format="PNG")
            raw = out.getvalue()
            _img_cache[url] = raw
            return raw
        except Exception as ex:
            logger.warning(f"Image fetch failed for {url}: {ex}")
            _img_cache[url] = None
            return None

    wb = Workbook()
    ws = wb.active
    ws.title = view_name[:31]   # Excel sheet name limit

    # ── Styles ────────────────────────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark navy
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
    ROW_FILL1 = PatternFill("solid", fgColor="FFFFFF")   # white
    ROW_FILL2 = PatternFill("solid", fgColor="EBF3FB")   # light blue
    CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=False)
    LEFT      = Alignment(horizontal="left",   vertical="center")

    # ── Headers ───────────────────────────────────────────────────────────────
    display_headers = [
        "Ad Preview" if c == "image_url" else c
        for c in df.columns
        if c != "ad_id"           # hide internal ad_id column
    ]
    visible_cols = [c for c in df.columns if c != "ad_id"]

    _PCT_HEADER = {c: f"{c} (%)" for c in PCT_METRICS}

    img_col_idx = None
    for idx, c in enumerate(visible_cols, start=1):
        cell = ws.cell(row=1, column=idx,
                       value="Ad Preview" if c == "image_url" else _PCT_HEADER.get(c, c))
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        if c == "image_url":
            img_col_idx = idx

    ws.row_dimensions[1].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────────
    IMG_ROW_H  = 90
    IMG_COL_W  = 18
    NUMBER_COLS = {"Spend", "Revenue", "CPM", "CPC", "CPT", "ROAS",
                   "CTR", "CVR", "CPC", "Impressions", "Clicks",
                   "Purchases", "Add to Cart", "Landing Page Views"}

    if img_col_idx:
        ws.column_dimensions[get_column_letter(img_col_idx)].width = IMG_COL_W

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        fill = ROW_FILL1 if row_idx % 2 == 0 else ROW_FILL2

        col_offset = 0
        for c_idx, col in enumerate(visible_cols, start=1):
            val = row.get(col)
            if col == "image_url":
                val = ""   # image anchored separately
            elif pd.isna(val):
                val = ""
            cell = ws.cell(row=row_idx, column=c_idx, value=val)
            cell.fill = fill
            cell.alignment = CENTER if col in NUMBER_COLS else LEFT

        # Embed image
        if has_images and img_col_idx:
            ws.row_dimensions[row_idx].height = IMG_ROW_H
            url = str(row.get("image_url", ""))
            img_bytes = _fetch(url)
            if img_bytes:
                xl_img = XLImage(io.BytesIO(img_bytes))
                xl_img.width  = 100
                xl_img.height = 80
                ws.add_image(xl_img, f"{get_column_letter(img_col_idx)}{row_idx}")
            elif url.startswith("http"):
                ws.cell(row=row_idx, column=img_col_idx).value = url

    # ── Auto column widths ────────────────────────────────────────────────────
    for c_idx, col in enumerate(visible_cols, start=1):
        if col == "image_url":
            continue
        header_len = len("Ad Preview" if col == "image_url" else col)
        data_lens  = [
            len(str(row.get(col, "")))
            for _, row in df.iterrows()
            if not pd.isna(row.get(col))
        ]
        best = max([header_len] + data_lens) if data_lens else header_len
        ws.column_dimensions[get_column_letter(c_idx)].width = min(best + 3, 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Google Sheets writer ──────────────────────────────────────────────────────

def write_custom_view_to_sheets(
    df: pd.DataFrame,
    tab_name: str,
    column_renames: dict | None = None,
    extra_cols: list[dict] | None = None,
) -> bool:
    """
    Write a custom view DataFrame to a named Google Sheets tab.
    Creates the tab if it doesn't exist; clears and rewrites if it does.

    column_renames : {old_name: new_name} applied after PCT conversion.
    extra_cols     : [{"name": "SUM of Amount spent (INR)", "source": "Spend"}, ...]
                     Appends a copy of 'source' column under 'name' at the end.
    Returns True on success.
    """
    import time
    import gspread
    from google.oauth2.service_account import Credentials
    from config import CREDENTIALS_PATH, SHEET_ID

    _SCOPES = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    _CHUNK_ROWS = 2000

    try:
        creds = Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=_SCOPES)
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_key(SHEET_ID)

        # Get or create worksheet
        try:
            worksheet = spreadsheet.worksheet(tab_name)
        except gspread.exceptions.WorksheetNotFound:
            worksheet = spreadsheet.add_worksheet(
                title=tab_name,
                rows=max(len(df) + 10, 100),
                cols=max(len(df.columns) + 5, 20),
            )

        # Convert df to list-of-lists — keep numerics as numbers so Sheets
        # stores them in Number format, not as text.
        import numpy as np
        export = df.copy()
        for col in export.columns:
            if pd.api.types.is_datetime64_any_dtype(export[col]):
                export[col] = export[col].dt.strftime("%Y-%m-%d")
        # Apply percentage conversion (decimals → ×100) for CTR/CVR/C2V Ratio
        for _pct_col in PCT_METRICS:
            if _pct_col in export.columns:
                export[_pct_col] = pd.to_numeric(export[_pct_col], errors="coerce") * 100
        # Rename percentage headers
        export = export.rename(columns={c: f"{c} (%)" for c in PCT_METRICS if c in export.columns})
        # Drop only ad_id (internal) — keep image_url as a clickable URL column in Sheets
        export = export.drop(columns=[c for c in ("ad_id",) if c in export.columns])

        # Append duplicate columns (e.g. "SUM of Amount spent (INR)" as a copy of Spend)
        if extra_cols:
            for _ec in extra_cols:
                _src = _ec.get("source", "")
                _nm  = _ec.get("name", "")
                if _src in export.columns and _nm:
                    export[_nm] = export[_src]

        # Apply custom column renames (e.g. Spend → Spends, Clicks → Link clicks)
        if column_renames:
            export = export.rename(columns=column_renames)

        # Per-column: numerics → fillna(0), text → fillna("").astype(str)
        for col in export.columns:
            if pd.api.types.is_numeric_dtype(export[col]):
                export[col] = export[col].fillna(0)
            else:
                export[col] = export[col].fillna("").astype(str)

        _MAX_CELL = 45000  # Google Sheets hard limit is 50,000 chars per cell

        def _native(v):
            if isinstance(v, np.integer): return int(v)
            if isinstance(v, np.floating): return float(v)
            # Truncate strings that exceed Sheets cell limit (causes 500 error)
            if isinstance(v, str) and len(v) > _MAX_CELL:
                return v[:_MAX_CELL]
            return v

        values = [export.columns.tolist()] + [
            [_native(v) for v in row] for row in export.values.tolist()
        ]

        worksheet.clear()
        time.sleep(0.5)

        # Use smaller chunks (500 rows) to avoid 500 payload-too-large errors
        _SAFE_CHUNK = 500
        for chunk_start in range(0, len(values), _SAFE_CHUNK):
            chunk = values[chunk_start: chunk_start + _SAFE_CHUNK]
            try:
                worksheet.update(f"A{chunk_start + 1}", chunk, value_input_option="USER_ENTERED")
            except Exception as _chunk_err:
                # If a chunk still fails, try row-by-row to isolate the bad cell
                logger.warning(f"Chunk at row {chunk_start + 1} failed ({_chunk_err}), retrying row-by-row…")
                for _ri, _row in enumerate(chunk[1:], start=chunk_start + 2):  # skip header re-send
                    try:
                        worksheet.update(f"A{_ri}", [_row], value_input_option="USER_ENTERED")
                    except Exception as _row_err:
                        logger.warning(f"  Row {_ri} skipped: {_row_err}")
            if chunk_start + _SAFE_CHUNK < len(values):
                time.sleep(1)

        logger.info(f"Custom view written to tab '{tab_name}' ({len(df)} rows).")
        return True

    except Exception as e:
        logger.error(f"Failed to write custom view to '{tab_name}': {e}")
        return False
