"""
Bundles aggregation results into the final response format:
  - plain_text  : human-readable summary
  - json_output : structured JSON (dataset, answer, method, table)
  - csv_bytes   : downloadable CSV bytes
  - excel_bytes : downloadable Excel with embedded ad images
"""

from __future__ import annotations

import io
import json
import logging

import pandas as pd

logger = logging.getLogger(__name__)


def _fmt(value) -> str:
    """Pretty-print a number."""
    if isinstance(value, float):
        return f"{value:,.2f}"
    if isinstance(value, int):
        return f"{value:,}"
    return str(value)


def build_plain_text(result: dict, user_question: str) -> str:
    """Generate a concise human-readable answer."""
    dataset  = result.get("dataset_used", "")
    method   = result.get("method", "")
    metrics  = result.get("metrics", {})
    val      = result.get("answer_value")
    table    = result.get("table")

    lines = []

    # ── Pincode count: give the direct answer first ────────────────────────────
    if "pincode_list" in result:
        unique_count = result.get("answer_value", 0)
        lines.append(f"**{unique_count} unique pincodes** were active in the selected period.")
        pin_list = result.get("pincode_list", [])
        if pin_list:
            lines.append(f"\n**Pincodes ({len(pin_list)}):** " + ", ".join(pin_list))
        lines.append(f"\n_Dataset: {dataset}_")
        lines.append(f"_Method: {method}_")
        return "\n".join(lines)

    # Primary answer
    if isinstance(val, (int, float)):
        metric_name = next(iter(metrics), "value")
        lines.append(f"**{metric_name}**: {_fmt(val)}")
    elif isinstance(val, dict):
        for k, v in val.items():
            if isinstance(v, (int, float)):
                lines.append(f"**{k}**: {_fmt(v)}")
    elif isinstance(val, pd.DataFrame):
        lines.append(f"Found **{len(val)}** rows in the breakdown.")

    # Summary metrics block
    if metrics:
        key_metrics = ["Spend", "Purchases", "Clicks", "Impressions",
                       "Pincode Days", "CTR", "CPC", "CPT", "CVR", "ROAS",
                       "Total Unique PCs"]
        shown = [(k, metrics[k]) for k in key_metrics if k in metrics]
        if shown:
            lines.append("\n**Summary:**")
            for k, v in shown:
                lines.append(f"  • {k}: {_fmt(v)}")

    lines.append(f"\n_Dataset: {dataset}_")
    lines.append(f"_Method: {method}_")
    return "\n".join(lines)


def build_json_output(result: dict, user_question: str) -> dict:
    """Structured JSON response."""
    table = result.get("table")
    table_records = []
    if isinstance(table, pd.DataFrame) and not table.empty:
        table_records = table.to_dict(orient="records")
        # Convert Timestamps to strings for JSON serialisation
        for row in table_records:
            for k, v in row.items():
                if hasattr(v, "isoformat"):
                    row[k] = v.isoformat()

    return {
        "dataset_used":   result.get("dataset_used", ""),
        "answer":         str(result.get("answer_value", "")),
        "method":         result.get("method", ""),
        "metrics":        {k: v for k, v in result.get("metrics", {}).items()
                           if isinstance(v, (int, float))},
        "table":          table_records,
        "download_ready": len(table_records) > 0,
    }


def build_csv_bytes(result: dict) -> bytes | None:
    """Return CSV bytes for download, or None if no table."""
    table = result.get("table")
    if not isinstance(table, pd.DataFrame) or table.empty:
        return None
    buf = io.StringIO()
    table.to_csv(buf, index=False)
    return buf.getvalue().encode("utf-8")


def build_excel_bytes(result: dict) -> bytes | None:
    """
    Return Excel (.xlsx) bytes with embedded ad images.

    If the table has an 'image_url' column, each unique URL is downloaded
    and embedded as an actual image in the 'Ad Preview' column.
    Falls back to writing the URL as text if the image can't be fetched.
    """
    table = result.get("table")
    if not isinstance(table, pd.DataFrame) or table.empty:
        return None

    try:
        import requests
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter
        from PIL import Image as PILImage
    except ImportError as e:
        logger.warning(f"Excel image export unavailable — missing dependency: {e}")
        return None

    df = table.copy()
    has_images = "image_url" in df.columns

    # Cache downloaded images: url → raw PNG bytes (not BytesIO — openpyxl reads lazily)
    _img_cache: dict[str, bytes | None] = {}

    def _fetch_image(url: str) -> bytes | None:
        """Download, resize to 120×120, return PNG bytes. None on any failure."""
        if not url or not str(url).startswith("http"):
            return None
        if url in _img_cache:
            return _img_cache[url]
        try:
            resp = requests.get(url, timeout=8)
            resp.raise_for_status()
            img = PILImage.open(io.BytesIO(resp.content))
            img.thumbnail((120, 120), PILImage.LANCZOS)
            out = io.BytesIO()
            # Always save as PNG — avoids format-detection issues with JPEG EXIF
            img.save(out, format="PNG")
            raw = out.getvalue()   # extract bytes BEFORE closing
            _img_cache[url] = raw
            return raw
        except Exception as ex:
            logger.warning(f"Could not fetch image {url}: {ex}")
            _img_cache[url] = None
            return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Result"

    # ── Write header ──────────────────────────────────────────────────────────
    # Replace 'image_url' header with 'Ad Preview'
    headers = []
    img_col_idx = None
    for i, col in enumerate(df.columns, start=1):
        if col == "image_url":
            headers.append("Ad Preview")
            img_col_idx = i
        else:
            headers.append(col)
    ws.append(headers)

    # Bold header row
    from openpyxl.styles import Font, Alignment
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # ── Row height & column widths ────────────────────────────────────────────
    IMG_ROW_HEIGHT = 90  # points — roughly 120px
    IMG_COL_WIDTH  = 18  # characters

    if img_col_idx:
        ws.column_dimensions[get_column_letter(img_col_idx)].width = IMG_COL_WIDTH

    # ── Write data rows ───────────────────────────────────────────────────────
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        row_data = []
        for col in df.columns:
            val = row[col]
            if col == "image_url":
                row_data.append("")          # placeholder; image anchored below
            else:
                row_data.append("" if pd.isna(val) else val)
        ws.append(row_data)

        if has_images and img_col_idx:
            ws.row_dimensions[row_idx].height = IMG_ROW_HEIGHT
            url = str(row.get("image_url", ""))
            img_bytes = _fetch_image(url)
            if img_bytes:
                # Give openpyxl a FRESH BytesIO each time — it reads lazily during save
                xl_img = XLImage(io.BytesIO(img_bytes))
                xl_img.width  = 100
                xl_img.height = 80
                cell_addr = f"{get_column_letter(img_col_idx)}{row_idx}"
                ws.add_image(xl_img, cell_addr)
            else:
                # Fallback: write the URL as text so user can open manually
                ws.cell(row=row_idx, column=img_col_idx).value = url

    # ── Auto-width non-image columns ──────────────────────────────────────────
    for i, col in enumerate(df.columns, start=1):
        if col == "image_url":
            continue
        col_letter = get_column_letter(i)
        lengths = [len(str(headers[i - 1]))] + [
            len(str(row[col])) for _, row in df.iterrows() if not pd.isna(row[col])
        ]
        max_len = max(lengths) if lengths else 10
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
